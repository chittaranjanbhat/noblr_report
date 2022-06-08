import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter


def design_xlsx(xlsx_file_path, sheet_name):

    wb = openpyxl.load_workbook(xlsx_file_path)

    # Name of the working sheet
    ws = wb[sheet_name]

    # Number of rows & columns
    n_row, n_col = ws.max_row, ws.max_column

    # Header Fill - Hex Value
    header_fill = PatternFill(patternType='solid', fgColor='386494')

    # Cell Fill - Hex Value
    # cell_fill = PatternFill(patternType='solid', fgColor='333333')

    # Header Font
    header_font = Font(bold=True, color='FFFFFF')

    # Cell Font
    cell_font = Font(color='365f96')

    for i in range(1, n_col + 1):
        for j in range(1, n_row + 1):
            if j == 1:
                ws.cell(row=j, column=i).fill = header_fill
                ws.cell(row=j, column=i).font = header_font
            else:
                # ws.cell(row=j, column=i).fill = cell_fill
                ws.cell(row=j, column=i).font = cell_font
            ws.cell(row=j, column=i).border = Border(top=Side(style='thin', color='000000'),
                                                     bottom=Side(style='thin', color='000000'),
                                                     left=Side(style='thin', color='000000'),
                                                     right=Side(style='thin', color='000000'))

    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].auto_size = True
    wb.save(xlsx_file_path)
